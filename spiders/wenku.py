# -*- coding: utf-8 -*-
import scrapy
import re
from openpyxl import Workbook
import openpyxl
import requests
from PIL import Image
from io import BytesIO
import os

class ZimukuSpider(scrapy.Spider):
    name = 'jingdongcollector'
    allowed_domains = ['list.jd.com','item.jd.com']
    start_urls = ["https://list.jd.com/list.html?cat=9192,12632,12633&page=1&sort=sort_totalsales15_desc&trans=1&JL=6_0_0"]
    medicine_re = re.compile('//item.jd.com/\d+.html')
    #https://list.jd.com
    page_re = re.compile('/list.html#cat=9192,12632,12633&page=\d+&sort=sort%5Ftotalsales15%5Fdesc&trans=1&JL=6_0_0')
    zh_pattern = re.compile(u'[\u4e00-\u9fa5]+')
    image_re = re.compile('//img\d+.360buyimg.com/n5/jfs/t\d+/\d+/\d+/\d+/[0-9a-zA-Z]+/[0-9a-zA-Z]+.[jpg|png]')
    guoyaozhunzi = []
    current_page = 1
    index = 1
    file_name = 'ganmao.xlsx'
    tag_dict = {'库内是否按批号管理':1,'库内是否按供应商管理':2,'产品名称':3,
                '包装大小':4,'药品分类':5,'药品商品名':6,
                '药品通用名':7,'批准文号':8,'生产企业':9,
                '品牌':10,'药品类型':11,'剂型':12,
                '适用人群':13,'产品规格':14,'用法用量':15,
                '有效期':16,'适用症/功能主治':17,}
    def parse(self, response):#只爬问答记录和往后翻页
        medicines = []
        page = ''
        next_page = 0
        #print(response.selector.xpath('//div[@id="J_main"]/div[@class="m-list"]//@href').extract())
        for url in response.selector.xpath('//div[@id="J_main"]/div[@class="m-list"]//@href').extract():
            url = url.replace('?','#')
            isMedicine=self.medicine_re.match(url)
            isPage=self.page_re.match(url)
            if isMedicine:
                medicines.append(url)
            elif isPage:
                page_num = int(re.search('page=\d+',url).group()[5:])
                #print('page num: {}'.format(page_num))
                if  page_num == self.current_page+1:
                    page = url
                    next_page = page_num
        #print(medicines)
        for url in medicines:
            #time.sleep(2000)
            yield scrapy.Request('https:'+url, callback=self.parse_medicine)
        print('current page is {}'.format(self.current_page))
        print('next page is {}'.format(next_page))
        self.current_page += 1            
        page = page.replace('#','?')
        #time.sleep(1000)
        yield scrapy.Request('https://list.jd.com'+page, callback=self.parse)
#这个funtion明天做        
    def parse_medicine(self, response):
        imagess = []
        imagesm = []
        imagesl = []
        image_url = response.selector.xpath('//div[@class="product-intro clearfix"]/div[@class="preview-wrap"]/div[@id="preview"]/div[@class="spec-list"]/div[@id="spec-list"]/ul[@class="lh"]//@src').extract()
        for url in image_url:
            isImage = self.image_re.match(url)
            if isImage:
                url = 'https:'+url
                imagess.append(url)
                imagesm.append(re.sub('/n\d+/','/n1/',url))
                imagesl.append(re.sub('/n\d+/','/n12/',url))
        text_content = response.selector.xpath('//div[@id="detail"]/div[@class="tab-con"]/div[@class="hide"]/div[@class="Ptable"]/div[@class="Ptable-item"]/dl/*/text()').extract()
        test = text_content[text_content.index('批准文号')+1].lstrip().rstrip()
        print(test)
        test = re.search('[0-9A-Z]+',test).group()
        if test in self.guoyaozhunzi:
            pass
        else:
            current_tag = ''
            pizhun = ''
            fold_name = './ganmao'
            if not os.path.exists(fold_name):
                os.mkdir(fold_name)
                os.mkdir(fold_name+'/images')
            file_name_download = fold_name+'/'+self.file_name  
            wb = Workbook()
            if os.path.exists(file_name_download):
                wb = openpyxl.load_workbook(filename=file_name_download)
                ws = wb["ganmaoyao"]
            else:
                ws = wb.create_sheet('ganmaoyao',0)
            for i in range(len(text_content)):
                if i%2 == 0:
                    current_tag = text_content[i].lstrip().rstrip()
                else:
                    if current_tag == '批准文号':
                        pizhun = re.search('[0-9A-Z]+',text_content[i].lstrip().rstrip()).group()
                        self.guoyaozhunzi.append(pizhun)
                    ws.cell(row=self.index, column=self.tag_dict[current_tag], value=text_content[i].lstrip().rstrip())    
            fold_name = './ganmao/images/'+pizhun
            os.mkdir(fold_name)
            fold_name += '/'
            string = ''
            if len(imagess)+len(imagesm)+len(imagesl) == 0:
                print(response.url)
            for image in imagess:
                response = requests.get(image)
                ima = Image.open(BytesIO(response.content))
                filename = fold_name+'s_'+image.split('/')[-1]
                ima.save(filename)
                #urllib.request.urlretrieve(image,filename=fold_name+image.split('/')[-1])
                string += filename+','
            string = string[:-1]
            ws.cell(row=self.index, column=18, value=string)
            string = ''
            for image in imagesm:
                response = requests.get(image)
                ima = Image.open(BytesIO(response.content))
                filename = fold_name+'m_'+image.split('/')[-1]
                ima.save(filename)
                #urllib.request.urlretrieve(image,filename=fold_name+image.split('/')[-1])
                string += filename+','
            string = string[:-1]
            ws.cell(row=self.index, column=19, value=string)
            string = ''
            for image in imagesl:
                response = requests.get(image)
                ima = Image.open(BytesIO(response.content))
                filename = fold_name+'l_'+image.split('/')[-1]
                ima.save(filename)
                #urllib.request.urlretrieve(image,filename=fold_name+image.split('/')[-1])
                string += filename+','
            string = string[:-1]
            ws.cell(row=self.index, column=20, value=string)
            self.index += 1
            wb.save(file_name_download)